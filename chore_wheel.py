#!/usr/bin/env python3


import pandas as pd
import random
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# --- CONFIGURATION ---
EXCEL_FILE = "ChoreAssignments.xlsx"
EMPLOYEE_COLUMN = "Employee"
OUT_COLUMN = "Out"
HISTORY_SHEET = "Chore History"
ASSIGNMENT_SHEET = "Assignments"

# Chore: number of people
WEEKLY_CHORES = {
    "Annex Bathrooms Upstairs/Downstairs": 3,
    "HCRU Upstairs Bathrooms": 3,
    "HCRU Downstairs Bathrooms": 3,
    "HCRU Upstairs Hallways/Stairs": 3,
    "HCRU Downstairs Hallways": 3,
    "Annex hallways/stairs": 3,
    "Conference Room": 2,
    "Microwaves HCRU": 1,
    "Annex Microwaves": 1,
    "Annex Headhouse": 2,
    "HCRU Headhouse": 2,
    "Growth Chamber Center Room": 2,
    "Growth Chamber Bathroom": 2,
}

# Chore week: Chore: number of people
MONTHLY_CHORES = {
    0: {"Growth Chamber Shower": 4},
    2: {"Fridges": 3},
}
# ----------------------


def load_excel(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} does not exist.")

    with pd.ExcelFile(file_path) as xls:
        sheet_names = xls.sheet_names

        # Load Chore History sheet if it exists
        try:
            history_df = pd.read_excel(xls, sheet_name=HISTORY_SHEET).fillna(0)
        except ValueError:
            history_df = pd.DataFrame(columns=[EMPLOYEE_COLUMN])

        # Load Assignments sheet if it exists
        if ASSIGNMENT_SHEET in sheet_names:
            main_df = pd.read_excel(xls, sheet_name=ASSIGNMENT_SHEET)
        else:
            # Look for a sheet with an Employee column to base the new Assignments sheet on
            for name in sheet_names:
                temp_df = pd.read_excel(xls, sheet_name=name)
                if EMPLOYEE_COLUMN in temp_df.columns:
                    main_df = temp_df.copy()  # Preserve all columns
                    print(
                        f"'{ASSIGNMENT_SHEET}' sheet not found. Created from sheet '{name}' with all columns preserved."
                    )
                    break
                else:
                    raise ValueError(
                        f"No sheet with a '{EMPLOYEE_COLUMN}' column found to create '{ASSIGNMENT_SHEET}'."
                    )

    return main_df, history_df


def get_week_and_month_counts(df):
    week_cols = [col for col in df.columns if col.startswith("Week")]
    month_cols = [col for col in df.columns if col.startswith("Month")]
    return len(week_cols), len(month_cols)


def get_last_chore_assignments(df):
    cols = [
        col for col in df.columns if col.startswith("Week") or col.startswith("Month")
    ]
    if not cols:
        return {}
    last_col = cols[-1]
    return dict(zip(df[EMPLOYEE_COLUMN], df[last_col]))


def get_available_people(df):
    if OUT_COLUMN in df.columns:
        return df[df[OUT_COLUMN] != True][EMPLOYEE_COLUMN].tolist()
    return df[EMPLOYEE_COLUMN].tolist()


def assign_chores_fairly(df, chores, excluded_people, last_assignments, history_df):
    available = [p for p in get_available_people(df) if p not in excluded_people]
    random.shuffle(available)
    assignments = {}
    used = set()

    for chore, needed in chores.items():
        eligible = []
        for person in available:
            if last_assignments.get(person) == chore or person in used:
                continue
            # Get chore count or default to 0
            if (
                chore in history_df.columns
                and person in history_df[EMPLOYEE_COLUMN].values
            ):
                count = history_df.loc[
                    history_df[EMPLOYEE_COLUMN] == person, chore
                ].values[0]
            else:
                count = 0
            eligible.append((person, count))
        eligible.sort(key=lambda x: x[1])

        selected = []
        for person, _ in eligible:
            if len(selected) < needed:
                selected.append(person)
                used.add(person)

        if len(selected) < needed:
            fallback = [p for p in available if p not in used]
            random.shuffle(fallback)
            for p in fallback:
                if len(selected) < needed:
                    selected.append(p)
                    used.add(p)

        for person in selected:
            assignments[person] = chore

    return assignments


def write_assignments(df, assignments, col_name):
    if col_name not in df.columns:
        df[col_name] = None
    for person, chore in assignments.items():
        df.loc[df[EMPLOYEE_COLUMN] == person, col_name] = chore
    return df


def update_history(history_df, assignments):
    for person, chore in assignments.items():
        if person not in history_df[EMPLOYEE_COLUMN].values:
            history_df = pd.concat(
                [history_df, pd.DataFrame([{EMPLOYEE_COLUMN: person}])],
                ignore_index=True,
            )
        if chore not in history_df.columns:
            history_df[chore] = 0
        history_df.loc[history_df[EMPLOYEE_COLUMN] == person, chore] += 1
    return history_df


def autofit_column_widths(file_path, sheet_names):
    wb = load_workbook(file_path)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)


def main(file_path=EXCEL_FILE, prompt=True):
    os.chdir(os.path.dirname(__file__))
    main_df, history_df = load_excel(file_path)
    week_count, month_count = get_week_and_month_counts(main_df)
    next_week_col = f"Week {week_count + 1}"
    last_assignments = get_last_chore_assignments(main_df)
    week_number = week_count + 1
    monthly_chores_this_week = {}
    if week_number > 2:
        monthly_chores_this_week = MONTHLY_CHORES.get((week_number + 1) % 4, {})

    monthly_assignments = {}
    excluded = []
    if monthly_chores_this_week:
        next_month_col = f"Month {month_count + 1}"
        monthly_assignments = assign_chores_fairly(
            main_df, monthly_chores_this_week, excluded, last_assignments, history_df
        )
        excluded += list(monthly_assignments.keys())
        main_df = write_assignments(main_df, monthly_assignments, next_month_col)
        history_df = update_history(history_df, monthly_assignments)
        print(f"Assigned monthly chores: {next_month_col}")

    weekly_assignments = assign_chores_fairly(
        main_df, WEEKLY_CHORES, excluded, last_assignments, history_df
    )
    all_assignments = {**monthly_assignments, **weekly_assignments}
    main_df = write_assignments(main_df, all_assignments, next_week_col)
    history_df = update_history(history_df, weekly_assignments)
    print(f"Assigned weekly chores: {next_week_col}")

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
        main_df.to_excel(writer, index=False, sheet_name=ASSIGNMENT_SHEET)
        history_df.to_excel(writer, index=False, sheet_name=HISTORY_SHEET)

    # Autofit columns in both sheets
    autofit_column_widths(file_path, [ASSIGNMENT_SHEET, HISTORY_SHEET])

    print(f"All chores and history saved to {file_path}.")
    if prompt:
        input("Done. Press ENTER to exit.")


if __name__ == "__main__":
    main(EXCEL_FILE, True)
